#! /usr/bin/python3
# converterExcelParaCsv.py - converte o os arquivos .xlsx de uma pasta para .csv

import csv
import sys
from easygui import *
import openpyxl
import os

def converterExcelParaCsv():

    # abre dialogbox para seleção do folder
    folder = diropenbox(msg="Selecione a pasta com os arquivos a converter")
    if folder is None:  # se clicou 'cancel'
        print("Cancel clicked.")
        sys.exit()
    os.chdir(folder)  # altera o diretório de trabalho para a pasta 'folder'

    os.makedirs('ExcelParaCsv', exist_ok=True)
    os.chdir('./ExcelParaCsv')

    for excelFile in os.listdir(folder):  # verifica se o arquivo é .xlsx
        if not excelFile.endswith('.xlsx'):
            continue

        wb = openpyxl.load_workbook(folder + '/' + excelFile)
        for sheetName in wb.sheetnames:
            # Percorre as planilhas do workbook em um loop
            sheet = wb[sheetName]
            print('Convertendo arquivo ' + excelFile + '...')

            # cria o nome do arquivo de destino a partir do nome original e do título da planilha
            csvFileName = excelFile.split('.')[0] + '-' + sheet.title + '.csv'
            outputFile = open(csvFileName, 'w', newline='')
            outputWriter = csv.writer(outputFile)

            # Percorre todas as linhas e fecha o arquivo de destinp
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []
                for columnNum in range(1, sheet.max_column + 1):
                    rowData.append(sheet.cell(row=rowNum, column=columnNum).value)
                    print(rowNum)
                outputWriter.writerow(rowData)
            outputFile.close()


if __name__ == '__main__': # executa se chamado diretamente
    converterExcelParaCsv()