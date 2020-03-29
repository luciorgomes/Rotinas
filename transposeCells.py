#! /usr/bin/python3
# transposeCells.py - Transpõe o conteúdo de linha em coluna (e vice-versa)

import openpyxl
import os
from easygui import *
import sys

def transposeCells():
    arq_input = fileopenbox(msg='Selecione o arquivo a ser transposto', filetypes=['*.xlsx'])
    if arq_input is None:  # se clicou 'cancel'
        print("Cancel clicked.")
        sys.exit()
    wb = openpyxl.load_workbook(arq_input)
    choices = wb.get_sheet_names()
    plan = choicebox(msg='Selecione a planilha a ser transposta:', choices=choices)
    if plan is None:  # se clicou 'cancel'
        print("Cancel clicked.")
        sys.exit()
    sheet = wb.get_sheet_by_name(plan)
    newWb = openpyxl.Workbook()
    newSheet = newWb.active
    for i in range (1, sheet.max_row + 1):
        for j in range (1, sheet.max_column + 1):
            newSheet.cell(row=j, column=i).value = sheet.cell(row=i, column=j).value
    newWb.save(os.path.dirname(arq_input) + '/transp_' + os.path.basename(arq_input))
    msgbox(msg='Salvo como '+ os.path.dirname(arq_input) + '/transp_' + os.path.basename(arq_input))

if __name__ == '__main__': # executa se chamado diretamente
    transposeCells()